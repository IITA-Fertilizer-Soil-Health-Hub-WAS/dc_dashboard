"""Coordinator-only "Plot anchor" ODK micro-form.

The country coordinator captures each trial's farmer-field anchor *in the field*
(governance: the anchor is ground-truthed in person, not office-pinned). Field GPS
may be offline, so capture rides a small ODK form the coordinator fills in ODK
Collect. This module:

  * generates that XLSForm on demand — one ``select_one`` of the project's
    still-unanchored trials plus a ``geopoint`` — as .xlsx bytes;
  * publishes it to the project's collection server via the shared publisher;
  * folds captured points back onto the CollectionUnits through the same
    containment-gated ``capture_anchor`` the in-app path uses.

See project memory: plot-election governance.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field

from openpyxl import Workbook

from apps.fieldwork.models import CandidatePlot

ANCHOR_FORM_TITLE = "Plot anchor (coordinator)"


def anchor_form_id(project) -> str:
    slug = project.code.lower().replace(" ", "_")
    return f"{slug}_plot_anchor"


def pending_anchor_trials(project) -> list[tuple[str, str]]:
    """(trial_key, candidate_ref) for every elected plot still awaiting its field
    anchor — the choice list the coordinator picks from while standing in the field."""
    rows = (
        CandidatePlot.objects.filter(
            project=project, status=CandidatePlot.Status.ELECTED
        )
        .select_related("collection_unit")
        .order_by("trial_key")
    )
    return [
        (c.trial_key, c.candidate_ref)
        for c in rows
        if c.collection_unit is not None and not c.collection_unit.anchor_captured
    ]


def build_anchor_xlsform(project) -> bytes:
    """Generate the coordinator anchor micro-form as XLSForm (.xlsx) bytes.

    The choice list is the project's currently-unanchored trials, so publishing a
    fresh copy after more elections widens the picker automatically."""
    from django.utils import timezone

    trials = pending_anchor_trials(project)
    wb = Workbook()

    survey = wb.active
    survey.title = "survey"
    survey.append(["type", "name", "label", "required", "hint"])
    survey.append([
        "note", "intro",
        f"Plot anchor capture — {project.name}. Stand on the farmer field, inside "
        "the elected plot, before capturing.", "", "",
    ])
    survey.append([
        "select_one trial", "trial_key", "Trial / plot to anchor", "yes",
        "Pick the trial you are standing on.",
    ])
    survey.append([
        "geopoint", "farmer_field", "Capture the farmer-field GPS", "yes",
        "Walk to the exact trial point, then capture.",
    ])

    choices = wb.create_sheet("choices")
    choices.append(["list_name", "name", "label"])
    for trial_key, ref in trials:
        choices.append(["trial", trial_key, f"{trial_key} (plot {ref})"])
    if not trials:
        # A select with no choices is an invalid XLSForm; leave a placeholder.
        choices.append(["trial", "__none__", "No trials awaiting anchor"])

    settings = wb.create_sheet("settings")
    settings.append(["form_title", "form_id", "version"])
    settings.append([
        ANCHOR_FORM_TITLE, anchor_form_id(project),
        timezone.now().strftime("%Y%m%d%H%M"),
    ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def publish_anchor_form(project):
    """Build the current anchor micro-form and publish it to the project's server.
    Returns (FormDefinition | None, PublishResult) from the shared publisher."""
    from apps.ingestion.publishing import publish_xlsform
    from apps.projects.models import FormDefinition

    xlsx = build_anchor_xlsform(project)
    return publish_xlsform(
        project, xlsx,
        filename=f"{anchor_form_id(project)}.xlsx",
        role=FormDefinition.Role.EXTRA,
        title=ANCHOR_FORM_TITLE,
    )


# --- Ingest → capture --------------------------------------------------------

@dataclass
class AnchorSyncStats:
    processed: int = 0
    captured: int = 0
    outside: int = 0
    skipped: int = 0
    errors: list = field(default_factory=list)


def _parse_geopoint(val) -> tuple[float | None, float | None]:
    """ODK geopoint is 'lat lon altitude accuracy' — take the first two."""
    if not isinstance(val, str):
        return None, None
    parts = val.split()
    try:
        return float(parts[0]), float(parts[1])
    except (ValueError, IndexError):
        return None, None


def anchor_form_for(project):
    """The most recently published anchor form for a project, if any."""
    from apps.projects.models import FormDefinition

    return (
        FormDefinition.objects.filter(project=project, title=ANCHOR_FORM_TITLE)
        .order_by("-published_at", "-created_at")
        .first()
    )


def apply_anchor_submissions(project, user=None) -> AnchorSyncStats:
    """Fold captured anchor submissions onto their trials' CollectionUnits.

    Reads the project's anchor-form submissions, pulls (trial_key, geopoint) from
    each, and runs the containment-gated capture. Idempotent: re-capturing the same
    point is harmless. A point that falls outside its elected plot is counted under
    `outside` and left uncaptured (the flag the containment gate is there to raise)."""
    from apps.fieldwork.anchor import capture_anchor
    from apps.fieldwork.models import CollectionUnit
    from apps.submissions.models import Submission

    stats = AnchorSyncStats()
    form = anchor_form_for(project)
    if form is None:
        return stats
    for sub in Submission.objects.filter(project=project, form=form).order_by("created_at"):
        payload = sub.raw_payload or {}
        trial_key = payload.get("trial_key") or payload.get("trial")
        lat, lon = _parse_geopoint(payload.get("farmer_field"))
        if not trial_key or lat is None:
            stats.skipped += 1
            continue
        unit = CollectionUnit.objects.filter(project=project, code=trial_key).first()
        if unit is None:
            stats.skipped += 1
            continue
        stats.processed += 1
        ok, msg = capture_anchor(user, unit, lat, lon)
        if ok:
            stats.captured += 1
        elif "outside" in msg.lower():
            stats.outside += 1
        else:
            stats.errors.append(f"{trial_key}: {msg}")
    return stats
