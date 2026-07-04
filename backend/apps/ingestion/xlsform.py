"""Generate an XLSForm (.xlsx) from a structured form spec.

XLSForm is the lingua franca every collection backend understands (ODK Central /
ONA / Kobo all run pyxform server-side on publish), so the in-app form builder
authors a *spec* and this turns it into the .xlsx that the existing publish rail
(apps.ingestion.publishing.publish_xlsform) pushes to the server. Keeping the
spec → XLSForm step here means we never talk a server-specific form API.

Spec shape (all keys optional except questions)::

    {
      "settings": {"form_title": "...", "form_id": "my_form", "version": "1"},
      "questions": [
        {"type": "text", "name": "farmer", "label": "Farmer name", "required": true},
        {"type": "integer", "name": "age", "label": "Age", "constraint": ". >= 0"},
        {"type": "select_one", "name": "crop", "label": "Crop", "list": "crop",
         "required": true},
        {"type": "begin_group", "name": "plot", "label": "Plot"},
        {"type": "end_group"},
      ],
      "choices": {"crop": [{"name": "maize", "label": "Maize"}]}
    }
"""
from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import Workbook

# XLSForm question types that reference a choice list.
_SELECT_TYPES = {"select_one", "select_multiple"}
# Structural rows that open/close a section (no label/name validation needed).
_STRUCTURE = {"end_group", "end_repeat"}

SURVEY_COLS = ["type", "name", "label", "required", "relevant", "constraint",
               "constraint_message", "hint", "calculation", "appearance"]
CHOICES_COLS = ["list_name", "name", "label"]


class XlsFormError(ValueError):
    """Raised when a spec can't produce a valid XLSForm."""


def slug(value: str) -> str:
    """A valid XForm identifier: letters/digits/underscore, not starting with a digit."""
    s = re.sub(r"[^0-9a-zA-Z_]+", "_", (value or "").strip()).strip("_").lower()
    if not s:
        s = "field"
    if s[0].isdigit():
        s = "_" + s
    return s


def _row_type(q: dict[str, Any]) -> str:
    t = (q.get("type") or "text").strip()
    if t in _SELECT_TYPES:
        raw = (q.get("list") or q.get("choices") or "").strip()
        if not raw:
            raise XlsFormError(f"'{q.get('name')}' is a {t} but has no choice list.")
        return f"{t} {slug(raw)}"
    if t in ("begin_group", "begin group"):
        return "begin group"
    if t in ("end_group", "end group"):
        return "end group"
    if t in ("begin_repeat", "begin repeat"):
        return "begin repeat"
    if t in ("end_repeat", "end repeat"):
        return "end repeat"
    return t


def build_xlsform(spec: dict[str, Any]) -> bytes:
    """Render a spec dict to XLSForm .xlsx bytes. Raises XlsFormError on bad input."""
    questions = spec.get("questions") or []
    if not questions:
        raise XlsFormError("A form needs at least one question.")

    wb = Workbook()
    survey = wb.active
    survey.title = "survey"
    survey.append(SURVEY_COLS)

    seen: set[str] = set()
    depth = 0
    for q in questions:
        rtype = q.get("type") or "text"
        row_type = _row_type(q)
        is_close = rtype in _STRUCTURE or rtype in ("end group", "end repeat")
        if is_close:
            depth -= 1
            survey.append([row_type] + [""] * (len(SURVEY_COLS) - 1))
            continue

        name = slug(q.get("name") or q.get("label") or "")
        if name in seen:
            raise XlsFormError(f"Duplicate field name '{name}'.")
        seen.add(name)
        is_open = row_type in ("begin group", "begin repeat")
        if is_open:
            depth += 1
        survey.append([
            row_type,
            name,
            q.get("label") or q.get("name") or name,
            "yes" if q.get("required") else "",
            q.get("relevant") or "",
            q.get("constraint") or "",
            q.get("constraint_message") or "",
            q.get("hint") or "",
            q.get("calculation") or "",
            q.get("appearance") or "",
        ])
    if depth != 0:
        raise XlsFormError("Unbalanced group/repeat — every begin needs an end.")

    choices = wb.create_sheet("choices")
    choices.append(CHOICES_COLS)
    for list_name, options in (spec.get("choices") or {}).items():
        ln = slug(list_name)
        for opt in options:
            choices.append([ln, slug(opt.get("name") or opt.get("label")),
                            opt.get("label") or opt.get("name") or ""])

    settings_sheet = wb.create_sheet("settings")
    settings_sheet.append(["form_title", "form_id", "version"])
    s = spec.get("settings") or {}
    settings_sheet.append([
        s.get("form_title") or "Untitled form",
        slug(s.get("form_id") or s.get("form_title") or "form"),
        str(s.get("version") or "1"),
    ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
