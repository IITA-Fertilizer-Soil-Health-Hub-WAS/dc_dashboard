"""Coordinator plot-anchor micro-form: generation, ingest→capture, and the
auto-dispatch + plot-ready notification that ride election / anchor capture."""
from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from apps.fieldwork.anchor import capture_anchor
from apps.fieldwork.anchor_form import (
    ANCHOR_FORM_TITLE,
    apply_anchor_submissions,
    build_anchor_xlsform,
    pending_anchor_trials,
)
from apps.fieldwork.dispatch import REGISTRATION_JOB_NAME, registration_job
from apps.fieldwork.election import elect_candidate
from apps.fieldwork.models import CandidatePlot, Job, UnitAssignment
from apps.projects.models import FormDefinition, Organization, Project
from apps.submissions.models import Submission

pytestmark = pytest.mark.django_db

# 0.01°-square around (lat -1.29, lon 36.80). GeoJSON coords are [lon, lat].
SQUARE = {"type": "Polygon", "coordinates": [[
    [36.80, -1.29], [36.81, -1.29], [36.81, -1.28], [36.80, -1.28], [36.80, -1.29]]]}


@pytest.fixture
def world():
    org = Organization.objects.create(code="o", name="O")
    uc = Project.objects.create(code="PROJ-A", name="A", organization=org)
    FormDefinition.objects.create(project=uc, ona_form_id=1,
                                  role=FormDefinition.Role.HH_REG)
    cand = CandidatePlot.objects.create(
        project=uc, trial_key="T1", candidate_ref="A", rank=1, geometry=SQUARE,
        centroid_lat=-1.285, centroid_lon=36.805)
    unit = elect_candidate(None, cand)
    return {"uc": uc, "cand": cand, "unit": unit}


# --- Feature 2: auto-dispatch + notify ---------------------------------------

def test_election_dispatches_registration_job(world):
    job = registration_job(world["uc"])
    assert job.name == REGISTRATION_JOB_NAME
    assert job.status == Job.Status.ACTIVE
    assert job.form is not None and job.form.role == FormDefinition.Role.HH_REG
    # The elected unit was folded into the job.
    assert UnitAssignment.objects.filter(job=job, unit=world["unit"]).exists()


def test_dispatch_is_idempotent(world):
    # Re-electing the same trial must not create a second job or duplicate assignment.
    elect_candidate(None, world["cand"])
    assert Job.objects.filter(project=world["uc"], name=REGISTRATION_JOB_NAME).count() == 1
    assert UnitAssignment.objects.filter(unit=world["unit"]).count() == 1


def test_anchor_capture_notifies_assigned_enumerator(world, django_user_model, mailoutbox):
    enum = django_user_model.objects.create_user(
        "e@x.org", "pw", is_active=True, full_name="Ed")
    job = registration_job(world["uc"])
    UnitAssignment.objects.filter(job=job, unit=world["unit"]).update(enumerator=enum)
    ok, _ = capture_anchor(None, world["unit"], -1.285, 36.805)
    assert ok
    assert len(mailoutbox) == 1
    assert "e@x.org" in mailoutbox[0].to
    assert world["unit"].code in mailoutbox[0].subject


def test_no_notification_when_unassigned(world, mailoutbox):
    ok, _ = capture_anchor(None, world["unit"], -1.285, 36.805)
    assert ok and len(mailoutbox) == 0


# --- Feature 1: anchor micro-form --------------------------------------------

def test_pending_trials_lists_unanchored_only(world):
    assert pending_anchor_trials(world["uc"]) == [("T1", "A")]
    capture_anchor(None, world["unit"], -1.285, 36.805)
    assert pending_anchor_trials(world["uc"]) == []  # now anchored → gone


def test_build_xlsform_has_trial_choice_and_geopoint(world):
    xlsx = build_anchor_xlsform(world["uc"])
    wb = load_workbook(io.BytesIO(xlsx))
    assert {"survey", "choices", "settings"} <= set(wb.sheetnames)
    survey_types = [row[0] for row in wb["survey"].iter_rows(min_row=2, values_only=True)]
    assert "select_one trial" in survey_types and "geopoint" in survey_types
    choice_names = [row[1] for row in wb["choices"].iter_rows(min_row=2, values_only=True)]
    assert "T1" in choice_names
    settings = list(wb["settings"].iter_rows(min_row=2, values_only=True))[0]
    assert settings[0] == ANCHOR_FORM_TITLE


def _anchor_form(uc):
    return FormDefinition.objects.create(
        project=uc, ona_form_id=99, role=FormDefinition.Role.EXTRA,
        title=ANCHOR_FORM_TITLE)


def _anchor_sub(uc, form, trial, geopoint, uuid):
    return Submission.objects.create(
        project=uc, form=form, ona_uuid=uuid, content_hash=uuid,
        raw_payload={"trial_key": trial, "farmer_field": geopoint})


def test_apply_anchor_captures_inside_and_flags_outside(world):
    uc = world["uc"]
    form = _anchor_form(uc)
    _anchor_sub(uc, form, "T1", "-1.285 36.805 0 5", "ok")       # inside the square
    stats = apply_anchor_submissions(uc)
    assert stats.captured == 1 and stats.outside == 0
    world["unit"].refresh_from_db()
    assert world["unit"].anchor_captured is True
    assert float(world["unit"].lat) == -1.285

    # A second submission outside the boundary is counted, unit unchanged.
    _anchor_sub(uc, form, "T1", "-1.35 36.805 0 5", "bad")        # south of the square
    stats2 = apply_anchor_submissions(uc)
    assert stats2.outside == 1


def test_apply_anchor_skips_unknown_trial_and_bad_gps(world):
    uc = world["uc"]
    form = _anchor_form(uc)
    _anchor_sub(uc, form, "T-NOPE", "-1.285 36.805 0 5", "unknown")  # no such unit
    _anchor_sub(uc, form, "T1", "", "nogps")                          # empty geopoint
    stats = apply_anchor_submissions(uc)
    assert stats.captured == 0 and stats.skipped == 2
