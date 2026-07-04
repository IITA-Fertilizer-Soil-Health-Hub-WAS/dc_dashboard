"""Spec → XLSForm generation (apps.ingestion.xlsform.build_xlsform)."""
from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from apps.ingestion.xlsform import XlsFormError, build_xlsform, slug


def _sheets(data: bytes):
    wb = load_workbook(io.BytesIO(data))
    out = {}
    for name in wb.sheetnames:
        out[name] = [[c for c in row] for row in wb[name].iter_rows(values_only=True)]
    return out


def test_slug_makes_valid_identifiers():
    assert slug("Farmer name") == "farmer_name"
    assert slug("2nd visit") == "_2nd_visit"      # can't start with a digit
    assert slug("  spaces  &  symbols! ") == "spaces_symbols"
    assert slug("") == "field"


def test_build_basic_form():
    data = build_xlsform({
        "settings": {"form_title": "Maize Trial", "form_id": "maize_trial", "version": "3"},
        "questions": [
            {"type": "text", "name": "Farmer name", "label": "Farmer name", "required": True},
            {"type": "integer", "name": "age", "label": "Age", "constraint": ". >= 0"},
        ],
    })
    sheets = _sheets(data)
    assert sheets["survey"][0][:3] == ["type", "name", "label"]
    # name slugified, required -> yes, constraint carried.
    assert sheets["survey"][1][:4] == ["text", "farmer_name", "Farmer name", "yes"]
    assert sheets["survey"][2][0] == "integer"
    assert ". >= 0" in sheets["survey"][2]
    # settings row
    assert sheets["settings"][1][:3] == ["Maize Trial", "maize_trial", "3"]


def test_select_one_references_choice_list():
    data = build_xlsform({
        "questions": [{"type": "select_one", "name": "crop", "label": "Crop", "list": "crop"}],
        "choices": {"crop": [{"name": "maize", "label": "Maize"},
                             {"name": "potato", "label": "Potato"}]},
    })
    sheets = _sheets(data)
    assert sheets["survey"][1][0] == "select_one crop"
    assert sheets["choices"][0] == ["list_name", "name", "label"]
    assert ["crop", "maize", "Maize"] in sheets["choices"]


def test_groups_and_repeats_round_trip():
    data = build_xlsform({
        "questions": [
            {"type": "begin_repeat", "name": "plots", "label": "Plots"},
            {"type": "geopoint", "name": "loc", "label": "Location"},
            {"type": "end_repeat"},
        ],
    })
    rows = [r[0] for r in _sheets(data)["survey"]]
    assert rows == ["type", "begin repeat", "geopoint", "end repeat"]


def test_select_without_list_errors():
    with pytest.raises(XlsFormError):
        build_xlsform({"questions": [{"type": "select_one", "name": "crop", "label": "Crop"}]})


def test_unbalanced_group_errors():
    with pytest.raises(XlsFormError):
        build_xlsform({"questions": [{"type": "begin_group", "name": "g", "label": "G"}]})


def test_duplicate_name_errors():
    with pytest.raises(XlsFormError):
        build_xlsform({"questions": [
            {"type": "text", "name": "a", "label": "A"},
            {"type": "text", "name": "a", "label": "A2"},
        ]})


def test_empty_form_errors():
    with pytest.raises(XlsFormError):
        build_xlsform({"questions": []})
